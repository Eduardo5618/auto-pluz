import os
import io
import unicodedata
from difflib import SequenceMatcher
from typing import Callable, Dict, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

try:
    from PIL import Image, ImageOps  # type: ignore
    _PIL_OK = True
except Exception:
    _PIL_OK = False


# Valores fijos (predeterminados para tus celdas)
IMG_WIDTH = 157   # 4.13 cm
IMG_HEIGHT = 210  # 5.55 cm

EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff"}

DESTINOS_DEF: Dict[str, List[str]] = {
    "RT_1": ["G3", "G5", "F3", "F5"],
    "RT_2": ["G3", "G5", "F3", "F5"],
}


OBJETIVOS_DEF: Dict[str, List[str]] = {
    "RT_1": ["carga primaria r", "carga primaria t", "carga secundaria r", "carga secundaria t"],
    "RT_2": ["carga primaria r", "carga primaria t", "carga secundaria r", "carga secundaria t"],
}

# ===== Helpers de similitud =====

def _normalize(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    for ch in (" ", "-", "_", ".", ","):
        s = s.replace(ch, "")
    return s

def _score(a: str, b: str) -> float:
    na, nb = _normalize(a), _normalize(b)
    if nb in na:
        return 1.0
    return SequenceMatcher(None, na, nb).ratio()

def _listar_imagenes(carpeta: str) -> List[str]:
    return sorted(
        os.path.join(carpeta, fn)
        for fn in os.listdir(carpeta)
        if os.path.splitext(fn.lower())[1] in EXTS
    )

# ===== Limpieza selectiva de imágenes =====

def _abrir_corrigiendo_exif(ruta: str):

    if not _PIL_OK:
        return None
    try:
        img = Image.open(ruta)
        return ImageOps.exif_transpose(img)
    except Exception:
        return None
    

def _anchor_cell_of(img) -> Optional[str]:
    try:
        col_idx = img.anchor._from.col + 1
        row_idx = img.anchor._from.row + 1
        return f"{get_column_letter(col_idx)}{row_idx}"
    except Exception:
        return None

def _limpiar_imagenes_en_celdas(ws, celdas: List[str]) -> None:
    keep = []
    target = set(celdas)
    for im in getattr(ws, "_images", []):
        if _anchor_cell_of(im) not in target:
            keep.append(im)
    ws._images = keep

# ===== Inserción (genérica por lista de rutas) =====
def _insertar_en_celdas(
        ws, 
        rutas_imgs: List[Optional[str]], 
        celdas: List[str],
        img_w: int, 
        img_h: int, 
        log: Callable[[str], None],
        autorotar: bool = True, 
) -> Tuple[int,int]:
    
    ok = fail = 0
    for ruta, celda in zip(rutas_imgs, celdas):
        try:
            if not ruta:
                log(f"⚠️ Sin coincidencia para {ws.title}:{celda}")
                continue
            if not os.path.isfile(ruta):
                raise FileNotFoundError(ruta)
            
            # Autorrotar y cargar correctamente en memoria si aplica
            if autorotar and _PIL_OK:
                pil_img = _abrir_corrigiendo_exif(ruta)
                if pil_img is not None:
                    buf = io.BytesIO()
                    pil_img.save(buf, format="PNG")  # convertir a PNG en memoria
                    buf.seek(0)
                    img = XLImage(buf)              # pasar BytesIO a openpyxl
                else:
                    img = XLImage(ruta)
            else:
                img = XLImage(ruta)
                
            img.width, img.height = img_w, img_h
            ws.add_image(img, celda)

            ok += 1
            log(f"✅ {os.path.basename(ruta)} → {ws.title}:{celda}")
        except Exception as e:
            fail += 1
            nombre = os.path.basename(ruta) if ruta else "—"
            log(f"❌ {nombre} → {ws.title}:{celda}: {e}")
    if len(rutas_imgs) < len(celdas):
        log(f"ℹ️ {ws.title}: faltaron {len(celdas) - len(rutas_imgs)} imagen(es).")
    return ok, fail

# ===== Selección por PATRÓN único (mantengo por compatibilidad) =====
def _mejores_coincidencias(carpeta: str, patron: str, k: int = 4, threshold: float = 0.45) -> List[str]:
    files = _listar_imagenes(carpeta)
    scores = [(_score(os.path.basename(p), patron), p) for p in files]
    scores.sort(key=lambda x: x[0], reverse=True)
    return [p for sc, p in scores if sc >= threshold][:k]

def procesar_fotos_por_patron(
    ruta_excel: str,
    carpeta_inicio: str,
    patron_inicio: str,
    carpeta_cierre: str,
    patron_cierre: str,
    destinos: Optional[Dict[str, List[str]]] = None,
    img_w: int = IMG_WIDTH,
    img_h: int = IMG_HEIGHT,
    limpiar_previas: bool = False,
    logger: Callable[[str], None] = print,
    autorotar: bool = True, 
) -> Dict[str, Dict[str, int]]:


    if destinos is None:
        destinos = DESTINOS_DEF

    if not os.path.isfile(ruta_excel):
        raise FileNotFoundError("Excel destino no existe.")
    if not os.path.isdir(carpeta_inicio):
        raise FileNotFoundError("Carpeta de INICIO inválida.")
    if not os.path.isdir(carpeta_cierre):
        raise FileNotFoundError("Carpeta de CIERRE inválida.")

    wb = load_workbook(ruta_excel)
    if "RT_1" not in wb.sheetnames: raise ValueError("No existe hoja 'RT_1'.")
    if "RT_2" not in wb.sheetnames: raise ValueError("No existe hoja 'RT_2'.")

    ws1, ws2 = wb["RT_1"], wb["RT_2"]
    dest1 = destinos.get("RT_1", [])[:4]
    dest2 = destinos.get("RT_2", [])[:4]

    if limpiar_previas:
        _limpiar_imagenes_en_celdas(ws1, dest1)
        _limpiar_imagenes_en_celdas(ws2, dest2)
        logger("🧹 Imágenes previas removidas en celdas destino.")

    logger(f"🔎 INICIO: buscando patrón “{patron_inicio}”…")
    sel_ini = _mejores_coincidencias(carpeta_inicio, patron_inicio, k=len(dest1))
    ok1, f1 = _insertar_en_celdas(ws1, sel_ini, dest1, img_w, img_h, logger, autorotar=autorotar)

    logger(f"🔎 CIERRE: buscando patrón “{patron_cierre}”…")
    sel_cie = _mejores_coincidencias(carpeta_cierre, patron_cierre, k=len(dest2))
    ok2, f2 = _insertar_en_celdas(ws2, sel_cie, dest2, img_w, img_h, logger, autorotar=autorotar)

    wb.save(ruta_excel)
    return {"RT_1": {"ok": ok1, "err": f1}, "RT_2": {"ok": ok2, "err": f2}}

# ===== Selección por OBJETIVOS (uno por celda) =====
def _seleccionar_por_objetivos(carpeta: str, objetivos: List[str], threshold: float = 0.45) -> List[Optional[str]]:
    """Para cada objetivo, elige el mejor archivo disponible (sin repetir dentro de la hoja)."""
    files = _listar_imagenes(carpeta)
    usados = set()
    seleccion: List[Optional[str]] = []

    for obj in objetivos:
        if not obj:
            seleccion.append(None)
            continue
        # rankear no usados
        candidatos = []
        for p in files:
            if p in usados:
                continue
            sc = _score(os.path.basename(p), obj)
            candidatos.append((sc, p))
        if not candidatos:
            seleccion.append(None); continue

        candidatos.sort(key=lambda x: x[0], reverse=True)
        best_sc, best_path = candidatos[0]
        if best_sc >= 0.45:
            usados.add(best_path)
            seleccion.append(best_path)
        else:
            seleccion.append(None)
    return seleccion

def procesar_fotos_por_objetivos(
    ruta_excel: str,
    carpeta_inicio: str,
    objetivos_inicio: List[str],
    carpeta_cierre: str,
    objetivos_cierre: List[str],
    destinos: Optional[Dict[str, List[str]]] = None,
    img_w: int = IMG_WIDTH,
    img_h: int = IMG_HEIGHT,
    limpiar_previas: bool = False,
    logger: Callable[[str], None] = print,
    autorotar: bool = True,
) -> Dict[str, Dict[str, int]]:
    """
    Mapea cada objetivo a una celda (misma posición) y coloca la mejor coincidencia
    por objetivo (sin repetir archivos dentro de la misma hoja).
    """
    if destinos is None:
        destinos = DESTINOS_DEF

    if not os.path.isfile(ruta_excel):
        raise FileNotFoundError("Excel destino no existe.")
    if not os.path.isdir(carpeta_inicio):
        raise FileNotFoundError("Carpeta de INICIO inválida.")
    if not os.path.isdir(carpeta_cierre):
        raise FileNotFoundError("Carpeta de CIERRE inválida.")

    wb = load_workbook(ruta_excel)
    if "RT_1" not in wb.sheetnames: raise ValueError("No existe hoja 'RT_1'.")
    if "RT_2" not in wb.sheetnames: raise ValueError("No existe hoja 'RT_2'.")
    ws1, ws2 = wb["RT_1"], wb["RT_2"]

    dest1 = (destinos.get("RT_1", []) or [])[:len(objetivos_inicio)]
    dest2 = (destinos.get("RT_2", []) or [])[:len(objetivos_cierre)]

    if limpiar_previas:
        _limpiar_imagenes_en_celdas(ws1, dest1)
        _limpiar_imagenes_en_celdas(ws2, dest2)
        logger("🧹 Imágenes previas removidas en celdas destino.")

    logger("🔎 INICIO: seleccionando por objetivos…")
    rutas_ini = _seleccionar_por_objetivos(carpeta_inicio, objetivos_inicio)
    ok1, f1 = _insertar_en_celdas(ws1, rutas_ini, dest1, img_w, img_h, logger, autorotar=autorotar)

    logger("🔎 CIERRE: seleccionando por objetivos…")
    rutas_cie = _seleccionar_por_objetivos(carpeta_cierre, objetivos_cierre)
    ok2, f2 = _insertar_en_celdas(ws2, rutas_cie, dest2, img_w, img_h, logger, autorotar=autorotar)

    wb.save(ruta_excel)
    return {"RT_1": {"ok": ok1, "err": f1}, "RT_2": {"ok": ok2, "err": f2}}

# ===== Wrapper con objetivos fijos (tu caso) =====
def procesar_fotos_predefinidos(
    ruta_excel: str,
    carpeta_inicio: str,
    carpeta_cierre: str,
    destinos: Optional[Dict[str, List[str]]] = None,
    img_w: int = IMG_WIDTH,
    img_h: int = IMG_HEIGHT,
    limpiar_previas: bool = False,
    logger: Callable[[str], None] = print,
    autorotar: bool = True,
) -> Dict[str, Dict[str, int]]:
    """
    Usa objetivos fijos:
    - carga primaria r, carga primaria t, carga secundaria r, carga secundaria t
    para INICIO (RT_1) y CIERRE (RT_2).
    """
    if destinos is None:
        destinos = DESTINOS_DEF

    return procesar_fotos_por_objetivos(
        ruta_excel=ruta_excel,
        carpeta_inicio=carpeta_inicio,
        objetivos_inicio=OBJETIVOS_DEF["RT_1"],
        carpeta_cierre=carpeta_cierre,
        objetivos_cierre=OBJETIVOS_DEF["RT_2"],
        destinos=destinos,
        img_w=img_w,
        img_h=img_h,
        limpiar_previas=limpiar_previas,
        logger=logger,
        autorotar=autorotar, 
    )