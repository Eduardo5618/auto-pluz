import re
import openpyxl
from openpyxl.utils import get_column_letter


DATE_PAT = re.compile(r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2})\b')


# ------------------------------
# Helpers
# ------------------------------
def _clean(s):
    if s is None: 
        return ""
    # Normaliza NBSP, mayúsculas, espacios y quita ":" final
    t = str(s).replace("\xa0", " ").upper().strip()
    t = re.sub(r"\s+", " ", t)
    if t.endswith(":"): 
        t = t[:-1]
    return t

def _resolver_ruta(path_like):
    if isinstance(path_like, (list, tuple)):
        if not path_like:
            raise FileNotFoundError("No se recibió ninguna ruta.")
        return str(path_like[0])
    return str(path_like)

def _leer_celda_robusto(ws, row, col):
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if val is None:
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr:
                return ws.cell(row=mr.min_row, column=mr.min_col).value
    return val

# ------------------------------
# FECHA / SED
# ------------------------------
def _buscar_valor_derecha_robusto(ws, etiqueta_variantes=("FECHA","FECHA:", " FECHA ", "FECHA "), max_offset=3):
    objetivos = {_clean(x) for x in etiqueta_variantes}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if _clean(cell.value) in objetivos:
                # prueba 1,2,... max_offset celdas a la derecha
                for k in range(1, max_offset+1):
                    val = _leer_celda_robusto(ws, cell.row, cell.column + k)
                    if val not in (None, ""):
                        return val, (cell.row, cell.column + k)
    return None, None

def _extraer_fecha_por_regex(ws):
    # Recorre toda la hoja buscando un texto que parezca fecha
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is None: continue
            s = str(v)
            m = DATE_PAT.search(s)
            if m:
                return m.group(0)
    return None


def encontrar_valor_a_la_derecha(ws, etiqueta_base):
    val, coord = _buscar_valor_derecha_robusto(ws, (etiqueta_base, f"{etiqueta_base}:", f" {etiqueta_base} ", f"{etiqueta_base} "))
    return val

# ------------------------------
# Detección de tabla
# ------------------------------

def encontrar_tabla_resumen(ws):
    max_row, max_col = ws.max_row, ws.max_column
    for r in range(1, max_row+1):
        mapa = {}
        for c in range(1, max_col+1):
            v = _clean(ws.cell(row=r, column=c).value)
            if v:
                mapa[v] = c
        aliases = {
            "LECTURA 1": {"LECTURA 1","LECTURA1","LECTURA I"},
            "LECTURA 2": {"LECTURA 2","LECTURA2","LECTURA II"}
        }
        ok = ("NOMBRE" in mapa and "FACTOR" in mapa and
              any(x in mapa for x in aliases["LECTURA 1"]) and
              any(x in mapa for x in aliases["LECTURA 2"]) and
              ("KWH" in mapa or "KW H" in mapa))
        if ok:
            colmap = {
                "NOMBRE": mapa["NOMBRE"],
                "FACTOR": mapa["FACTOR"],
                "LECTURA 1": next(mapa[x] for x in aliases["LECTURA 1"] if x in mapa),
                "LECTURA 2": next(mapa[x] for x in aliases["LECTURA 2"] if x in mapa),
                "KWH": mapa.get("KWH", mapa.get("KW H")),
            }
            return {"row": r, "colmap": colmap}
    return None

def encontrar_tabla_por_totalizador(ws):
    objetivos = {"TOTALIZADOR", "TOTALIZADOR - FACTURADOS"}
    max_row, max_col = ws.max_row, ws.max_column
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            v = _clean(ws.cell(row=r, column=c).value)
            if v in objetivos:
                hdr_row = r - 1
                if hdr_row < 1:
                    continue
                colmap = {
                    "NOMBRE":    c,
                    "FACTOR":    c + 1,
                    "LECTURA 1": c + 2,
                    "LECTURA 2": c + 3,
                    "KWH":       c + 4,
                }
                return {"row": hdr_row, "colmap": colmap}
    return None


def leer_filas_tabla(ws, tabla_info, nombres_a_leer=("TOTALIZADOR","ALP1","ALP 1","ALP2","CLIENTES")):

    if not tabla_info:
        return {}

    r0 = tabla_info["row"]
    colmap = tabla_info["colmap"]
    targets = {_clean(n) for n in nombres_a_leer}

    def _leer_robusto(r, c):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val is None:
            for mr in ws.merged_cells.ranges:
                if cell.coordinate in mr:
                    return ws.cell(row=mr.min_row, column=mr.min_col).value
        return val

    resultados = {}
    r = r0 + 1
    max_row = ws.max_row

    consecutivos_vacios = 0
    VACIOS_TOPE = 10      
    BUSQUEDA_MAX = 200    

    exploradas = 0
    while r <= max_row and exploradas < BUSQUEDA_MAX:
        nombre_raw = _leer_robusto(r, colmap["NOMBRE"])
        nombre = _clean(nombre_raw)

        if not nombre:
            consecutivos_vacios += 1
            if consecutivos_vacios >= VACIOS_TOPE:
                break
            r += 1
            exploradas += 1
            continue

        consecutivos_vacios = 0

        if nombre.startswith("TOTALIZADOR - FACTURADOS"):
            break

        if any(nombre.startswith(t) for t in targets):
            fila = {
                "FACTOR":    _leer_robusto(r, colmap["FACTOR"]),
                "LECTURA 1": _leer_robusto(r, colmap["LECTURA 1"]),
                "LECTURA 2": _leer_robusto(r, colmap["LECTURA 2"]),
                "KWH":       _leer_robusto(r, colmap["KWH"]),
            }
            resultados[nombre] = fila

        r += 1
        exploradas += 1

    print(f"[DEBUG] Filas leídas (final): {resultados}")
    return resultados                      

def leer_por_ancla_totalizador(ws):
    import re
    pat = re.compile(r"^TOTALIZADOR\b")
    max_row, max_col = ws.max_row, ws.max_column

    def _leer(r, c):
        cell = ws.cell(row=r, column=c)
        v = cell.value
        if v is None:
            for mr in ws.merged_cells.ranges:
                if cell.coordinate in mr:
                    return ws.cell(row=mr.min_row, column=mr.min_col).value
        return v

    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            v = _clean(ws.cell(row=r, column=c).value)
            if pat.match(v):
                out = {}
                # TOTALIZADOR en r
                out["TOTALIZADOR"] = {
                    "FACTOR":    _leer(r, c+1),
                    "LECTURA 1": _leer(r, c+2),
                    "LECTURA 2": _leer(r, c+3),
                    "KWH":       _leer(r, c+4),
                }
                # ALP1 en r+1 (si existe)
                nombre_alp = _clean(_leer(r+1, c))
                if nombre_alp.startswith("ALP"):
                    out["ALP1"] = {
                        "FACTOR":    _leer(r+1, c+1),
                        "LECTURA 1": _leer(r+1, c+2),
                        "LECTURA 2": _leer(r+1, c+3),
                        "KWH":       _leer(r+1, c+4),
                    }
                print(f"[DEBUG] leer_por_ancla_totalizador -> {out}")
                return out
    return {}

# ------------------------------
# Pegado en output
# ------------------------------

def pegar_en_output(ruta_output, hoja_output, valores, mapeo_celdas, visible=False):

    import os
    import pythoncom
    import win32com.client as win32

    def _resolver_ruta(path_like):
        if isinstance(path_like, (list, tuple)):
            if not path_like:
                raise FileNotFoundError("No se recibió ninguna ruta.")
            return str(path_like[0])
        return str(path_like)

    ruta_output = _resolver_ruta(ruta_output)
    ruta_output_abs = os.path.abspath(ruta_output)

    # Inicia COM (clave si llamas desde hilos/GUI)
    need_uninit = False
    try:
        pythoncom.CoInitialize()
        need_uninit = True
    except pythoncom.com_error:
        pass

    app = None
    wb = None
    try:
        app = win32.DispatchEx("Excel.Application")
        app.Visible = bool(visible)
        app.DisplayAlerts = False

        wb = app.Workbooks.Open(ruta_output_abs)
        try:
            ws = wb.Worksheets(hoja_output) 
        except Exception as e:
            hojas = [sh.Name for sh in wb.Worksheets]
            raise KeyError(f"No existe hoja '{hoja_output}' en '{ruta_output_abs}'. Hojas: {hojas}") from e

        # ---- FECHA / SED ----
        if "fecha" in mapeo_celdas:
            celda_fecha = mapeo_celdas["fecha"]
            ws.Range(celda_fecha).Value = valores.get("fecha")
        if "sed" in mapeo_celdas:
            celda_sed = mapeo_celdas["sed"]
            ws.Range(celda_sed).Value = valores.get("sed")

        # ---- Tabla (TOTALIZADOR, ALP1, ...) ----
        filas = valores.get("filas", {})
        for nombre, campos in mapeo_celdas.items():
            if nombre in ("fecha", "sed"):
                continue
            if isinstance(campos, dict) and nombre in filas:
                for campo, celda in campos.items():
                    if campo in filas[nombre]:
                        ws.Range(celda).Value = filas[nombre][campo]

        wb.Save()  

    finally:
        if wb is not None:
            wb.Close(SaveChanges=True)
        if app is not None:
            app.Quit()

        try:
            import gc
            del wb; del app
            gc.collect()
        except:
            pass
        if need_uninit:
            try:
                pythoncom.CoUninitialize()
            except:
                pass


def encontrar_tabla_por_totalizador(ws):

    objetivos = {"TOTALIZADOR", "TOTALIZADOR - FACTURADOS"}
    max_row, max_col = ws.max_row, ws.max_column

    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            v = _clean(ws.cell(row=r, column=c).value)
            if v in objetivos:
                # headers una fila arriba del TOTALIZADOR
                hdr_row = r - 1
                if hdr_row < 1:
                    continue
                colmap = {
                    "NOMBRE":    c,
                    "FACTOR":    c + 1,
                    "LECTURA 1": c + 2,
                    "LECTURA 2": c + 3,
                    "KWH":       c + 4,
                }
                return {"row": hdr_row, "colmap": colmap}
    return None


def extraer_y_pegar(ruta_input, hoja_input, ruta_output, hoja_output, mapeo_celdas, debug=True):

    ruta_input  = _resolver_ruta(ruta_input)
    ruta_output = _resolver_ruta(ruta_output)

    wb_in = openpyxl.load_workbook(ruta_input, data_only=True)
    if hoja_input not in wb_in.sheetnames and hoja_input.upper() in wb_in.sheetnames:
        hoja_input = hoja_input.upper()
    if hoja_input not in wb_in.sheetnames:
        raise KeyError(f"No existe hoja '{hoja_input}' en '{ruta_input}'. Hojas: {wb_in.sheetnames}")
    ws_in = wb_in[hoja_input]

    # 1) FECHA y SED
    fecha = encontrar_valor_a_la_derecha(ws_in, "FECHA")
    sed   = encontrar_valor_a_la_derecha(ws_in, "SED")
    if fecha in (None, ""):
        # plan B si era fórmula sin cache
        wb_in_formula = openpyxl.load_workbook(ruta_input, data_only=False)
        fecha = _extraer_fecha_por_regex(wb_in_formula[hoja_input])

    # 2) Tabla
    tabla_info = encontrar_tabla_resumen(ws_in) or encontrar_tabla_por_totalizador(ws_in)
    filas = leer_filas_tabla(ws_in, tabla_info, nombres_a_leer=("TOTALIZADOR","ALP1","ALP 1","ALP2","CLIENTES"))
    if not filas:
        filas = leer_por_ancla_totalizador(ws_in)

    if debug:
        print(f"[DEBUG] Input: {ruta_input}")
        print(f"[DEBUG] Hoja: {hoja_input}")
        print(f"[DEBUG] FECHA={fecha!r}  SED={sed!r}")
        print(f"[DEBUG] Tabla headers: {tabla_info}")
        print(f"[DEBUG] Filas: {filas}")

    valores = {"fecha": fecha, "sed": sed, "filas": filas}
    pegar_en_output(ruta_output, hoja_output, valores, mapeo_celdas)

    if debug:
        print(f"[DEBUG] Pegado en '{ruta_output}' hoja '{hoja_output}' OK")